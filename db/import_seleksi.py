import sqlite3
from pathlib import Path
from openpyxl import load_workbook
from decimal import Decimal

DB = "seleksi.db"
SOURCE = Path("source")

FILES = [
    "data seleksi1.xlsx",
    "seleksi2.xlsx",
    "seleksi3.xlsx",
    "seleksi4.xlsx",
    "seleksi5.xlsx",
]

# Kamus pemetaan nama kolom Excel -> nama kolom Database SQLite
ALIASES = {
    "No.": "no",
    "Status": "status",
    "Pilihan": "pilihan",
    "Jalur Pendaftaran": "jalur_pendaftaran",
    "Nama Orang Tua": "nama_orang_tua",
    "Kontak Orang Tua": "kontak_orang_tua",
    "Nomor Pendaftaran": "nomor_pendaftaran",
    "NISN": "nisn",
    "NIK": "nik",
    "No. KK": "no_kk",
    "Nama": "nama",
    "Asal Sekolah": "asal_sekolah",
    "Pilihan 1": "pilihan_1",
    "Jarak Pilihan 1": "jarak_pilihan_1",
    "Pilihan 2": "pilihan_2",
    "Jarak Pilihan 2": "jarak_pilihan_2",
    "Pilihan 3": "pilihan_3",
    "Surat Hasil Diagnosa": "surat_hasil_diagnosa",
    "Terdaftar di DTKS": "terdaftar_di_dtks",
    "Terdaftar di Non DTKS": "terdaftar_di_non_dtks",
    "Kartu Program Penanggulangan Kemiskinan": "kartu_program_penanggulangan_kemiskinan",
    "Surat Tugas Orang Tua atau Surat Keterangan": "surat_tugas_orang_tua_atau_surat_keterangan",
    "Provinsi Perpindahan": "provinsi_perpindahan",
    "Kab/Kota Perpindahan": "kab_kota_perpindahan",
    "Kecamatan Perpindahan": "kecamatan_perpindahan",
    "Kelurahan Perpindahan": "kelurahan_perpindahan",
    "RT Perpindahan": "rt_perpindahan",
    "RW Perpindahan": "rw_perpindahan",
    "Alamat Lengkap Perpindahan": "alamat_lengkap_perpindahan",
    "Titik Koordinat Perpindahan": "titik_koordinat_perpindahan",
    "Surat Pemindahan Tugas": "surat_pemindahan_tugas",
    "Surat Tugas Mengajar/Bekerja": "surat_tugas_mengajar_bekerja",
    "Sertifikat Pendidik": "sertifikat_pendidik",
    "Prioritas": "prioritas",
    "Skor Kejuaraan": "skor_kejuaraan",
    "Skor Ujikom": "skor_ujikom",
    "Skor Akhir": "skor_akhir",
    "Penyelenggara": "penyelenggara",
    "Nama Prestasi": "nama_prestasi",
    "Kategori": "kategori",
    "Pelaksanaan": "pelaksanaan",
    "Prestasi": "prestasi",
    "Skor": "skor",
    "Jenis Kelamin": "jenis_kelamin",
    "Tempat Lahir": "tempat_lahir",
    "Tanggal Lahir": "tanggal_lahir",
    "Provinsi": "provinsi",
    "Kab/Kota": "kab_kota",
    "Kecamatan": "kecamatan",
    "Kelurahan": "kelurahan",
    "RT": "rt",
    "RW": "rw",
    "Alamat Lengkap": "alamat_lengkap",
    "Titik Koordinat": "titik_koordinat",
}

def clean(x):
    if x is None:
        return None
    if isinstance(x, (float, int)):
        try:
            return f"{Decimal(str(x)):.0f}".strip()
        except Exception:
            pass
    val_str = str(x).strip()
    if val_str.endswith(".0"):
        val_str = val_str[:-2]
    return val_str if val_str else None

def get_table_columns(cur, table_name):
    """Mendapatkan daftar kolom asli dari tabel SQLite."""
    try:
        cols = [c[1] for c in cur.execute(f"PRAGMA table_info({table_name})").fetchall()]
        return cols
    except Exception:
        return []

def find_header(ws):
    """Mencari baris header berdasarkan kolom 'No.' dan 'NIK'."""
    for row_num, row in enumerate(ws.iter_rows(values_only=True), 1):
        vals = [clean(x) for x in row]
        if "No." in vals and "NIK" in vals:
            return row_num, vals
    return None, None

def main():
    conn = sqlite3.connect(DB)
    cur = conn.cursor()

    total_inserted = 0

    for filename in FILES:
        path = SOURCE / filename

        if not path.exists():
            print(f"⚠️ File tidak ditemukan: {filename}")
            continue

        print("\n" + "=" * 60)
        print(f"📖 Membaca: {filename}")

        wb = load_workbook(path, read_only=True, data_only=True)

        for ws in wb.worksheets:
            # Nama tabel SQLite disesuaikan dengan nama sheet (lowercase & strip)
            table_name = ws.title.strip().lower().replace("-", "_").replace(" ", "_")

            # Cek ketersediaan tabel di DB
            db_cols = get_table_columns(cur, table_name)
            if not db_cols:
                # Bila nama sheet 'ofline', ganti penanganan ke sheet umum
                continue

            header_row, header = find_header(ws)
            if not header_row:
                continue

            print(f"📄 Sheet '{ws.title}' ➔ Tabel Database '{table_name}'")

            # Petakan kolom Excel ke kolom tabel SQLite
            mapping = {}
            for i, name in enumerate(header):
                if name in ALIASES and ALIASES[name] in db_cols:
                    mapping[ALIASES[name]] = i

            if not mapping:
                continue

            target_cols = list(mapping.keys())
            placeholders = ",".join(["?"] * len(target_cols))
            sql_insert = f"INSERT INTO {table_name} ({','.join(target_cols)}) VALUES ({placeholders})"

            sheet_count = 0

            for row in ws.iter_rows(min_row=header_row + 1, values_only=True):
                if not any(x is not None for x in row):
                    continue

                row_data = []
                for col in target_cols:
                    idx = mapping[col]
                    val = clean(row[idx]) if idx < len(row) else None
                    row_data.append(val)

                # Validasi baris: harus memiliki setidaknya salah satu identitas utama
                nik_idx = target_cols.index("nik") if "nik" in target_cols else None
                no_pemb_idx = target_cols.index("nomor_pendaftaran") if "nomor_pendaftaran" in target_cols else None
                nama_idx = target_cols.index("nama") if "nama" in target_cols else None

                has_id = False
                if nik_idx is not None and row_data[nik_idx]:
                    has_id = True
                elif no_pemb_idx is not None and row_data[no_pemb_idx]:
                    has_id = True
                elif nama_idx is not None and row_data[nama_idx]:
                    has_id = True

                if not has_id:
                    continue

                cur.execute(sql_insert, tuple(row_data))
                sheet_count += 1
                total_inserted += 1

            conn.commit()
            print(f"   ➕ Berhasil memasukkan: {sheet_count} baris")

        wb.close()

    conn.close()

    print("\n" + "=" * 60)
    print("✨ IMPOR SELESAI")
    print(f"🆕 Total Data Masuk: {total_inserted}")

if __name__ == "__main__":
    main()

